"""
EVI Daily Task Manager
======================
Backend engine: SQLite database + Excel report generator.

Usage (Terminal):
  python task_manager.py save              — read New Entry sheet → save to DB
  python task_manager.py recall_date       — prompt for date → export to Excel
  python task_manager.py recall_week       — prompt for year/week → export to Excel
  python task_manager.py all               — export every record to Excel
  python task_manager.py rebuild_excel     — rebuild full DailyTaskLog.xlsx from DB
"""

import sqlite3, os, sys, calendar
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_FILE    = os.path.join(BASE_DIR, "tasks.db")
EXCEL_FILE = os.path.join(BASE_DIR, "DailyTaskLog.xlsx")

# ── Palette ────────────────────────────────────────────────────
C_NAVY     = "1F3864"
C_BLUE     = "2E75B6"
C_LTBLUE   = "DEEAF1"
C_PALE     = "EBF3FB"
C_WHITE    = "FFFFFF"
C_LABEL_BG = "D6E4F7"
C_WARN     = "FFF2CC"
C_OK       = "E2EFDA"
C_BORDER   = "2E75B6"
C_TODAY_BG = "FF6600"
C_TODAY_FG = "FFFFFF"
C_WKND_BG = "F2D7D5"
C_CAL_HDR  = "2E75B6"

def bdr(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def bdr_thick():
    s = Side(style="medium", color=C_NAVY)
    return Border(top=s, bottom=s, left=s, right=s)

def hdr_cell(ws, row, col, value, width=None, bg=C_NAVY, fg=C_WHITE, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = bdr()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def dat_cell(ws, row, col, value="", even=True, h="center", wrap=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=C_LTBLUE if even else C_WHITE)
    c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    c.border = bdr()
    if fmt:
        c.number_format = fmt
    return c

# ─────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_date    TEXT NOT NULL,
                day_name      TEXT NOT NULL,
                description   TEXT,
                location      TEXT,
                dc_code       TEXT DEFAULT 'EVI01',
                add_info      TEXT,
                week_number   INTEGER,
                year          INTEGER,
                saved_at      TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        conn.commit()
    print(f"[DB] Ready: {DB_FILE}")

def save_record(entry_date, description, location, dc_code="EVI01", add_info=""):
    if isinstance(entry_date, str):
        entry_date = datetime.strptime(entry_date, "%Y-%m-%d").date()
    day_name    = entry_date.strftime("%A")
    week_number = int(entry_date.strftime("%W"))
    year        = entry_date.year
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO tasks (entry_date, day_name, description, location,
                               dc_code, add_info, week_number, year)
            VALUES (?,?,?,?,?,?,?,?)
        """, (entry_date.isoformat(), day_name, description, location,
              dc_code, add_info, week_number, year))
        conn.commit()
        return cur.lastrowid

def fetch_by_date(target_date):
    if isinstance(target_date, date):
        target_date = target_date.isoformat()
    with get_conn() as conn:
        return conn.execute(
            "SELECT * FROM tasks WHERE entry_date=? ORDER BY saved_at",
            (target_date,)
        ).fetchall()

def fetch_by_week(year, week):
    with get_conn() as conn:
        return conn.execute(
            "SELECT * FROM tasks WHERE year=? AND week_number=? ORDER BY entry_date, saved_at",
            (year, week)
        ).fetchall()

def fetch_all():
    with get_conn() as conn:
        return conn.execute(
            "SELECT * FROM tasks ORDER BY entry_date, saved_at"
        ).fetchall()

# ─────────────────────────────────────────────────────────────
# SHARED TABLE COLUMNS
# ─────────────────────────────────────────────────────────────
COLUMNS = [
    ("#",              5,  "center"),
    ("Date",          14,  "center"),
    ("Day",           12,  "center"),
    ("Description",   44,  "left"),
    ("Location",      13,  "center"),
    ("DC Code",       14,  "center"),
    ("Additional Information", 36, "left"),
    ("Saved At",      18,  "center"),
]

def write_db_rows(ws, rows, start_row):
    for ri, row in enumerate(rows, start=start_row):
        even = ri % 2 == 0
        ws.row_dimensions[ri].height = 22
        d = dict(row)
        vals = [d["id"], d["entry_date"], d["day_name"],
                d["description"] or "", d["location"] or "",
                d["dc_code"] or "EVI01", d["add_info"] or "", d["saved_at"] or ""]
        for ci, (val, (_, _, ha)) in enumerate(zip(vals, COLUMNS), 1):
            c = dat_cell(ws, ri, ci, val, even=even, h=ha, wrap=(ci in (4, 7)))
            if ci == 2:
                c.number_format = "DD-MMM-YYYY"

# ─────────────────────────────────────────────────────────────
# SHEET: NEW ENTRY  (single-row titles, inputs below)
# ─────────────────────────────────────────────────────────────
def build_entry_sheet(wb):
    ws = wb.create_sheet("New Entry", 0)

    # Column widths:  A=Date  B=Day  C=Description  D=Location  E=DC Code  F=Add.Info
    widths = {"A": 16, "B": 14, "C": 48, "D": 14, "E": 16, "F": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── ROW 1: Title banner ──────────────────────────────────
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "EVI DAILY TASK LOG  —  Datacenter EVI01"
    t.font = Font(name="Arial", bold=True, size=16, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.border = bdr()
    ws.row_dimensions[1].height = 38

    # ── ROW 2: Spacer ────────────────────────────────────────
    ws.row_dimensions[2].height = 6

    # ── ROW 3: Column titles (SINGLE ROW) ────────────────────
    titles = ["Date", "Day", "Description", "Location\n(DH1–DH6)", "DC Code",
              "Additional\nInformation"]
    ws.row_dimensions[3].height = 32
    for ci, title in enumerate(titles, 1):
        c = ws.cell(row=3, column=ci, value=title)
        c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr_thick()

    # ── ROW 4: Input fields (horizontal, one per title) ──────
    ws.row_dimensions[4].height = 30
    today = datetime.today().date()

    # A4 – Date (calendar date-picker via data-validation)
    c_date = ws.cell(row=4, column=1, value=today)
    c_date.number_format = "DD-MMM-YYYY"
    c_date.font = Font(name="Arial", size=12, bold=True)
    c_date.fill = PatternFill("solid", fgColor=C_PALE)
    c_date.alignment = Alignment(horizontal="center", vertical="center")
    c_date.border = bdr_thick()

    dv_date = DataValidation(
        type="date",
        operator="greaterThan",
        formula1="2020-01-01",
        allow_blank=False,
        showInputMessage=True,
        promptTitle="Select Date",
        prompt="Click this cell → a calendar date-picker will appear.\nSelect the date for this task entry.",
        showErrorMessage=True,
        errorTitle="Invalid Date",
        error="Please enter a valid date.",
        showDropDown=False,
    )
    ws.add_data_validation(dv_date)
    dv_date.add(ws["A4"])

    # B4 – Day (auto-calculated from date)
    c_day = ws.cell(row=4, column=2, value=today.strftime("%A"))
    c_day.font = Font(name="Arial", size=11, italic=True)
    c_day.fill = PatternFill("solid", fgColor=C_LABEL_BG)
    c_day.alignment = Alignment(horizontal="center", vertical="center")
    c_day.border = bdr_thick()

    # C4 – Description
    c_desc = ws.cell(row=4, column=3, value="")
    c_desc.font = Font(name="Arial", size=11)
    c_desc.fill = PatternFill("solid", fgColor=C_PALE)
    c_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c_desc.border = bdr_thick()

    # D4 – Location dropdown (DH1–DH6)
    c_loc = ws.cell(row=4, column=4, value="DH1")
    c_loc.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
    c_loc.fill = PatternFill("solid", fgColor=C_BLUE)
    c_loc.alignment = Alignment(horizontal="center", vertical="center")
    c_loc.border = bdr_thick()

    dv_loc = DataValidation(
        type="list",
        formula1='"DH1,DH2,DH3,DH4,DH5,DH6"',
        allow_blank=False,
        showDropDown=False,
        showInputMessage=True,
        promptTitle="Datahall",
        prompt="Select a Datahall location: DH1 through DH6",
    )
    ws.add_data_validation(dv_loc)
    dv_loc.add(ws["D4"])

    # E4 – DC Code (pre-filled)
    c_dc = ws.cell(row=4, column=5, value="EVI01")
    c_dc.font = Font(name="Arial", bold=True, size=11)
    c_dc.fill = PatternFill("solid", fgColor=C_PALE)
    c_dc.alignment = Alignment(horizontal="center", vertical="center")
    c_dc.border = bdr_thick()

    # F4 – Additional Information
    c_add = ws.cell(row=4, column=6, value="")
    c_add.font = Font(name="Arial", size=11)
    c_add.fill = PatternFill("solid", fgColor=C_PALE)
    c_add.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c_add.border = bdr_thick()

    # ── ROW 5: Spacer ────────────────────────────────────────
    ws.row_dimensions[5].height = 6

    # ── ROW 6: Tip bar ───────────────────────────────────────
    ws.merge_cells("A6:F6")
    tip = ws["A6"]
    tip.value = '  SAVE ▶  Run in Terminal:   python3 task_manager.py save       |       RECALL ▶  python3 task_manager.py recall_date   or   recall_week'
    tip.font = Font(name="Arial", size=9, bold=True, color="595959")
    tip.fill = PatternFill("solid", fgColor=C_WARN)
    tip.alignment = Alignment(horizontal="left", vertical="center")
    tip.border = bdr()
    ws.row_dimensions[6].height = 22

    # ── ROW 7: Spacer ────────────────────────────────────────
    ws.row_dimensions[7].height = 6

    # ── ROW 8: Click-hint for date ────────────────────────────
    ws.merge_cells("A8:F8")
    hint = ws["A8"]
    hint.value = "  TIP:  Click on the Date cell (A4) — Excel will show a calendar date-picker with today highlighted.   Day auto-fills from the date."
    hint.font = Font(name="Arial", size=9, italic=True, color="2E75B6")
    hint.fill = PatternFill("solid", fgColor=C_LTBLUE)
    hint.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[8].height = 20

    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110
    return ws

# ─────────────────────────────────────────────────────────────
# SHEET: CALENDAR  (visual monthly calendar, today highlighted)
# ─────────────────────────────────────────────────────────────
def build_calendar_sheet(wb):
    name = "Calendar"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    today = date.today()
    year  = today.year
    month = today.month

    # Column widths (cols B–H for Mon–Sun)
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 12

    # ── Title ────────────────────────────────────────────────
    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value = f"EVI DAILY TASK LOG — {today.strftime('%B %Y').upper()}"
    t.font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── Sub-title: Today indicator ────────────────────────────
    ws.merge_cells("B2:H2")
    ts = ws["B2"]
    ts.value = f"Today:  {today.strftime('%A, %d-%b-%Y')}      (Week {int(today.strftime('%W'))})"
    ts.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    ts.fill = PatternFill("solid", fgColor=C_TODAY_BG)
    ts.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # ── Day-name headers (row 3) ──────────────────────────────
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    ws.row_dimensions[3].height = 24
    for ci, dn in enumerate(day_names, 2):
        c = ws.cell(row=3, column=ci, value=dn)
        bg = C_WKND_BG if dn in ("Sat", "Sun") else C_CAL_HDR
        fg = C_NAVY if dn in ("Sat", "Sun") else C_WHITE
        c.font = Font(name="Arial", bold=True, size=11, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()

    # ── Calendar grid ─────────────────────────────────────────
    cal = calendar.Calendar(firstweekday=0)  # Monday first
    weeks = cal.monthdayscalendar(year, month)

    for wi, week in enumerate(weeks):
        row = 4 + wi
        ws.row_dimensions[row].height = 32
        for di, day_num in enumerate(week):
            col = 2 + di
            c = ws.cell(row=row, column=col)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr()

            if day_num == 0:
                # Empty cell (day from prev/next month)
                c.value = ""
                c.fill = PatternFill("solid", fgColor="F5F5F5")
                c.font = Font(name="Arial", size=10, color="CCCCCC")
            elif day_num == today.day:
                # TODAY — bright highlight
                c.value = day_num
                c.font = Font(name="Arial", bold=True, size=14, color=C_TODAY_FG)
                c.fill = PatternFill("solid", fgColor=C_TODAY_BG)
                c.border = bdr_thick()
            elif di >= 5:
                # Weekend
                c.value = day_num
                c.font = Font(name="Arial", size=11, color="943126")
                c.fill = PatternFill("solid", fgColor=C_WKND_BG)
            else:
                # Normal weekday
                c.value = day_num
                c.font = Font(name="Arial", size=11, color=C_NAVY)
                c.fill = PatternFill("solid", fgColor=C_WHITE)

    # ── Legend row ────────────────────────────────────────────
    legend_row = 4 + len(weeks) + 1
    ws.row_dimensions[legend_row].height = 20

    lg = ws.cell(row=legend_row, column=2, value="Today")
    lg.font = Font(name="Arial", bold=True, size=9, color=C_TODAY_FG)
    lg.fill = PatternFill("solid", fgColor=C_TODAY_BG)
    lg.alignment = Alignment(horizontal="center", vertical="center")
    lg.border = bdr()

    lw = ws.cell(row=legend_row, column=3, value="Weekend")
    lw.font = Font(name="Arial", size=9, color="943126")
    lw.fill = PatternFill("solid", fgColor=C_WKND_BG)
    lw.alignment = Alignment(horizontal="center", vertical="center")
    lw.border = bdr()

    ln = ws.cell(row=legend_row, column=4, value="Weekday")
    ln.font = Font(name="Arial", size=9, color=C_NAVY)
    ln.fill = PatternFill("solid", fgColor=C_WHITE)
    ln.alignment = Alignment(horizontal="center", vertical="center")
    ln.border = bdr()

    # ── Previous month mini-view ──────────────────────────────
    prev_row = legend_row + 2
    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    ws.merge_cells(f"B{prev_row}:D{prev_row}")
    pm = ws.cell(row=prev_row, column=2,
                 value=f"< {calendar.month_name[prev_month]} {prev_year}")
    pm.font = Font(name="Arial", bold=True, size=10, color=C_BLUE)
    pm.alignment = Alignment(horizontal="left", vertical="center")

    # ── Next month mini-view ──────────────────────────────────
    next_month = month + 1 if month < 12 else 1
    next_year  = year if month < 12 else year + 1
    ws.merge_cells(f"F{prev_row}:H{prev_row}")
    nm = ws.cell(row=prev_row, column=6,
                 value=f"{calendar.month_name[next_month]} {next_year} >")
    nm.font = Font(name="Arial", bold=True, size=10, color=C_BLUE)
    nm.alignment = Alignment(horizontal="right", vertical="center")

    ws.sheet_view.zoomScale = 110
    return ws

# ─────────────────────────────────────────────────────────────
# SHEET: ALL RECORDS
# ─────────────────────────────────────────────────────────────
def build_all_records_sheet(wb, rows):
    name = "All Records"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"ALL RECORDS — EVI Daily Task Log  ({len(rows)} entries)"
    t.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.row_dimensions[2].height = 28
    for ci, (name_, w, _) in enumerate(COLUMNS, 1):
        hdr_cell(ws, 2, ci, name_, width=w, bg=C_BLUE)

    write_db_rows(ws, rows, start_row=3)
    ws.freeze_panes = "A3"
    if rows:
        ws.auto_filter.ref = f"A2:H{2 + len(rows)}"
    ws.sheet_view.zoomScale = 100
    return ws

# ─────────────────────────────────────────────────────────────
# SHEET: RECALL – BY DATE
# ─────────────────────────────────────────────────────────────
def build_recall_date_sheet(wb, rows=None, target_date=None):
    name = "Recall - By Date"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "RECALL BY DATE"
    t.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    hdr_cell(ws, 3, 1, "Search Date", bg=C_LABEL_BG, fg=C_NAVY, size=10, width=14)
    sd = ws.cell(row=3, column=2,
                 value=target_date if target_date else datetime.today().date())
    sd.number_format = "DD-MMM-YYYY"
    sd.font = Font(name="Arial", size=11, bold=True)
    sd.fill = PatternFill("solid", fgColor=C_PALE)
    sd.alignment = Alignment(horizontal="center", vertical="center")
    sd.border = bdr()
    ws.row_dimensions[3].height = 26

    ws.merge_cells("A4:H4")
    tip = ws["A4"]
    tip.value = "  Run:  python3 task_manager.py recall_date   to refresh results"
    tip.font = Font(name="Arial", size=9, italic=True, color="595959")
    tip.fill = PatternFill("solid", fgColor=C_WARN)
    ws.row_dimensions[4].height = 18

    if rows and len(rows) > 0:
        ws.merge_cells("A5:H5")
        badge = ws.cell(row=5, column=1,
                        value=f"  {len(rows)} record(s) found for {target_date}")
        badge.font = Font(name="Arial", size=10, bold=True, color="375623")
        badge.fill = PatternFill("solid", fgColor=C_OK)
        ws.row_dimensions[5].height = 20

        ws.row_dimensions[6].height = 28
        for ci, (name_, w, _) in enumerate(COLUMNS, 1):
            hdr_cell(ws, 6, ci, name_, width=w, bg=C_BLUE)
        write_db_rows(ws, rows, start_row=7)
    else:
        ws.merge_cells("A6:H6")
        ws.cell(row=6, column=1, value="  No records found.").font = \
            Font(name="Arial", size=10, italic=True, color="595959")

    ws.freeze_panes = "A7"
    return ws

# ─────────────────────────────────────────────────────────────
# SHEET: RECALL – BY WEEK
# ─────────────────────────────────────────────────────────────
def build_recall_week_sheet(wb, rows=None, year=None, week=None):
    name = "Recall - By Week"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "RECALL BY WEEK"
    t.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    hdr_cell(ws, 3, 1, "Year", bg=C_LABEL_BG, fg=C_NAVY, size=10, width=10)
    ws.cell(row=3, column=2,
            value=year or datetime.today().year).font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=3, column=2).fill = PatternFill("solid", fgColor=C_PALE)
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=2).border = bdr()

    hdr_cell(ws, 3, 3, "Week #", bg=C_LABEL_BG, fg=C_NAVY, size=10, width=10)
    ws.cell(row=3, column=4,
            value=week or int(datetime.today().strftime("%W"))).font = \
        Font(name="Arial", size=11, bold=True)
    ws.cell(row=3, column=4).fill = PatternFill("solid", fgColor=C_PALE)
    ws.cell(row=3, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=4).border = bdr()
    ws.row_dimensions[3].height = 26

    ws.merge_cells("A4:H4")
    tip = ws["A4"]
    tip.value = "  Run:  python3 task_manager.py recall_week   to refresh results"
    tip.font = Font(name="Arial", size=9, italic=True, color="595959")
    tip.fill = PatternFill("solid", fgColor=C_WARN)
    ws.row_dimensions[4].height = 18

    if rows and len(rows) > 0:
        ws.merge_cells("A5:H5")
        badge = ws.cell(row=5, column=1,
                        value=f"  {len(rows)} record(s) — Year {year}, Week {week}")
        badge.font = Font(name="Arial", size=10, bold=True, color="375623")
        badge.fill = PatternFill("solid", fgColor=C_OK)
        ws.row_dimensions[5].height = 20

        ws.row_dimensions[6].height = 28
        for ci, (name_, w, _) in enumerate(COLUMNS, 1):
            hdr_cell(ws, 6, ci, name_, width=w, bg=C_BLUE)
        write_db_rows(ws, rows, start_row=7)
    else:
        ws.merge_cells("A6:H6")
        ws.cell(row=6, column=1, value="  No records found.").font = \
            Font(name="Arial", size=10, italic=True, color="595959")

    ws.freeze_panes = "A7"
    return ws

# ─────────────────────────────────────────────────────────────
# REBUILD ENTIRE WORKBOOK
# ─────────────────────────────────────────────────────────────
def rebuild_excel(recall_date=None, recall_week_args=None):
    wb = Workbook()
    wb.remove(wb.active)

    build_entry_sheet(wb)
    build_calendar_sheet(wb)
    build_all_records_sheet(wb, fetch_all())

    if recall_date:
        build_recall_date_sheet(wb, fetch_by_date(recall_date), target_date=recall_date)
    else:
        build_recall_date_sheet(wb, rows=None, target_date=datetime.today().date())

    if recall_week_args:
        yr, wk = recall_week_args
        build_recall_week_sheet(wb, fetch_by_week(yr, wk), year=yr, week=wk)
    else:
        build_recall_week_sheet(wb, rows=None,
                                year=datetime.today().year,
                                week=int(datetime.today().strftime("%W")))

    wb.active = wb["New Entry"]
    wb.save(EXCEL_FILE)
    print(f"[Excel] Saved: {EXCEL_FILE}")

# ─────────────────────────────────────────────────────────────
# READ ENTRY FROM EXCEL
# ─────────────────────────────────────────────────────────────
def read_entry_from_excel():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["New Entry"]
    entry_date  = ws["A4"].value
    location    = ws["D4"].value
    description = ws["C4"].value
    dc_code     = ws["E4"].value or "EVI01"
    add_info    = ws["F4"].value or ""

    if isinstance(entry_date, datetime):
        entry_date = entry_date.date()
    elif isinstance(entry_date, str):
        for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
            try:
                entry_date = datetime.strptime(entry_date, fmt).date()
                break
            except ValueError:
                continue
    return {
        "entry_date":  entry_date,
        "location":    location,
        "description": description,
        "dc_code":     dc_code,
        "add_info":    add_info,
    }

# ─────────────────────────────────────────────────────────────
# CLI COMMANDS
# ─────────────────────────────────────────────────────────────
def cmd_save():
    data = read_entry_from_excel()
    if not data["description"] or str(data["description"]).strip() in (
            "", "Enter task description here..."):
        print("[ERROR] Description is empty. Fill in the New Entry sheet first.")
        return
    if not data["location"]:
        print("[ERROR] Location is empty. Select DH1-DH6.")
        return
    row_id = save_record(
        entry_date=data["entry_date"], description=data["description"],
        location=data["location"], dc_code=data["dc_code"],
        add_info=data["add_info"],
    )
    print(f"[SAVED] Record #{row_id} — {data['entry_date']}  {data['location']}")
    rebuild_excel()
    os.system(f'open "{EXCEL_FILE}"')

def cmd_recall_date():
    raw = input("Enter date [DD-MMM-YYYY or YYYY-MM-DD, Enter=today]: ").strip()
    if not raw:
        target = datetime.today().date()
    else:
        for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                target = datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        else:
            print("[ERROR] Unrecognized date format.")
            return
    rows = fetch_by_date(target)
    print(f"[RECALL] {len(rows)} record(s) for {target}")
    rebuild_excel(recall_date=target.isoformat())
    os.system(f'open "{EXCEL_FILE}"')

def cmd_recall_week():
    yr_raw = input(f"Year [Enter={datetime.today().year}]: ").strip()
    wk_raw = input(f"Week # [Enter={int(datetime.today().strftime('%W'))}]: ").strip()
    year = int(yr_raw) if yr_raw else datetime.today().year
    week = int(wk_raw) if wk_raw else int(datetime.today().strftime("%W"))
    rows = fetch_by_week(year, week)
    print(f"[RECALL] {len(rows)} record(s) — Year {year}, Week {week}")
    rebuild_excel(recall_week_args=(year, week))
    os.system(f'open "{EXCEL_FILE}"')

def cmd_all():
    rows = fetch_all()
    print(f"[ALL] {len(rows)} total record(s)")
    rebuild_excel()
    os.system(f'open "{EXCEL_FILE}"')

if __name__ == "__main__":
    init_db()

    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
    if count == 0:
        print("[DB] Seeding sample entries...")
        save_record(date(2026, 4, 6), "Replaced faulty PDU breaker on rack A-14; tested all circuits OK.", "DH3", "EVI01", "Completed without downtime")
        save_record(date(2026, 4, 6), "Monthly visual inspection of racks B1-B20.", "DH1", "EVI01", "No issues found")
        save_record(date(2026, 4, 7), "Installed 2x new UPS units in Row C.", "DH5", "EVI01", "Requires firmware update next shift")
        save_record(date(2026, 4, 7), "Network cabling audit on DH2 top-of-rack switches.", "DH2", "EVI01", "3 loose cables re-seated")
        save_record(date(2026, 4, 8), "Cooling unit PM — cleaned filters, checked airflow.", "DH4", "EVI01", "All readings normal")

    cmd = sys.argv[1] if len(sys.argv) > 1 else "rebuild_excel"
    cmds = {"save": cmd_save, "recall_date": cmd_recall_date,
            "recall_week": cmd_recall_week, "all": cmd_all,
            "rebuild_excel": lambda: (rebuild_excel(), os.system(f'open "{EXCEL_FILE}"'))}
    cmds.get(cmd, lambda: print(__doc__))()
