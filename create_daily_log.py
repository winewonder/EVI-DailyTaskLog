from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

OUTPUT = "/Users/spatel/Documents/EVI/Tools/DailyTasks/DailyTaskLog.xlsx"

C_HDR_BG   = "1F3864"
C_HDR_FG   = "FFFFFF"
C_ROW_EVEN = "DEEAF1"
C_ROW_ODD  = "FFFFFF"
C_INPUT_BG = "EBF3FB"
C_BORDER   = "2E75B6"

def border():
    s = Side(style="thin", color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
    c.fill = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def data_cell(ws, row, col, value="", even=True, h_align="center", wrap=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=C_ROW_EVEN if even else C_ROW_ODD)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    c.border = border()
    if fmt:
        c.number_format = fmt
    return c

wb = Workbook()
ws = wb.active
ws.title = "Daily Task Log"

# ── Title banner ─────────────────────────────────────────────
ws.merge_cells("A1:G1")
title = ws["A1"]
title.value = "EVI DAILY TASK LOG — Datacenter EVI01"
title.font = Font(name="Arial", bold=True, size=14, color=C_HDR_FG)
title.fill = PatternFill("solid", fgColor=C_HDR_BG)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# ── Column headers (row 2) ────────────────────────────────────
#  Col: A=Date  B=Day  C=Description  D=Location  E=Datacenter Code  F=Additional Info
columns = [
    (1, "Date",             14),
    (2, "Day",              12),
    (3, "Description",      42),
    (4, "Location",         14),
    (5, "Datacenter Code",  18),
    (6, "Additional\nInformation", 36),
]

ws.row_dimensions[2].height = 30
for col, name, width in columns:
    header_cell(ws, 2, col, name, width)

# ── Data rows (3 to 52 → 50 blank input rows) ────────────────
dv_dh = DataValidation(
    type="list",
    formula1='"DH1,DH2,DH3,DH4,DH5,DH6"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Invalid location",
    error="Choose DH1 through DH6",
)
ws.add_data_validation(dv_dh)

# Sample entries
samples = [
    (datetime(2026, 4, 6).date(), "Replaced faulty PDU breaker on rack A-14; tested all circuits OK.", "DH3", "EVI01", "Completed without downtime"),
    (datetime(2026, 4, 6).date(), "Monthly visual inspection of racks B1-B20.", "DH1", "EVI01", "No issues found"),
    (datetime(2026, 4, 7).date(), "Installed 2x new UPS units in Row C.", "DH5", "EVI01", "Requires firmware update next shift"),
]

START_ROW = 3
TOTAL_ROWS = 52   # rows 3-52 = 50 data rows

for r in range(START_ROW, TOTAL_ROWS + 1):
    idx = r - START_ROW          # 0-based sample index
    even = (r % 2 == 0)
    ws.row_dimensions[r].height = 20

    if idx < len(samples):
        date_val, desc, loc, dc, addl = samples[idx]
        day_name = date_val.strftime("%A")
        data_cell(ws, r, 1, date_val,  even, "center", fmt="DD-MMM-YYYY")
        data_cell(ws, r, 2, day_name,  even, "center")
        data_cell(ws, r, 3, desc,      even, "left",   wrap=True)
        data_cell(ws, r, 4, loc,       even, "center")
        data_cell(ws, r, 5, dc,        even, "center")
        data_cell(ws, r, 6, addl,      even, "left",   wrap=True)
    else:
        # Blank input row – pre-fill Datacenter Code with EVI01
        for col in range(1, 7):
            bg = C_ROW_EVEN if even else C_ROW_ODD
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="center" if col != 3 else "left",
                vertical="center", wrap_text=(col in (3, 6))
            )
            c.border = border()
            if col == 1:
                c.number_format = "DD-MMM-YYYY"
            if col == 5:
                c.value = "EVI01"   # default datacenter code

    # Attach dropdown to Location column
    dv_dh.add(ws.cell(row=r, column=4))

# ── Freeze header rows ────────────────────────────────────────
ws.freeze_panes = "A3"

# ── Auto-filter ───────────────────────────────────────────────
ws.auto_filter.ref = f"A2:F{TOTAL_ROWS}"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
